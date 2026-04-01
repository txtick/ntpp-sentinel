#!/usr/bin/env python3
"""
GHL Contact Type Update Script — Final Version
Sets all 573 contacts to their correct type in one run.
Types: Active Customer, Past Customer, Lead, Internal, Vendor, Do Not Contact

USAGE:
    python ghl_update.py          # Dry run — preview only, no changes
    python ghl_update.py --live   # Apply all changes for real

SETUP:
    1. Fill in GHL_API_KEY below
    2. Keep GHL_Contact_Type_Update.xlsx in the same folder
    3. Always dry run first to confirm
"""

import requests
import time
import sys
from openpyxl import load_workbook
from collections import Counter

# ── CONFIGURATION ─────────────────────────────────────────────────────────────
GHL_API_KEY     = 'pit-7119c529-05e1-4e4d-bed8-a8b00ff6f0df'
GHL_LOCATION_ID = '74pWgJL459jf8GzoYOmz'
EXCEL_FILE    = 'GHL_Contact_Type_Update.xlsx'
DELAY_SECONDS = 0.25

BASE_URL = 'https://services.leadconnectorhq.com'
HEADERS  = {
    'Authorization': f'Bearer {GHL_API_KEY}',
    'Content-Type': 'application/json',
    'Version': '2021-07-28',
}

# Confirmed working API values
TYPE_MAP = {
    'Active Customer': 'customer',
    'Past Customer':   'past_customer',
    'Lead':            'lead',
    'Do Not Contact':  'do_not_contact',
    'Internal':        'internal',
    'Vendor':          'vendor',
}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def cv(v):
    s = str(v).strip() if v is not None else ''
    return '' if s in ['nan','None','—',''] else s

def read_contacts():
    wb = load_workbook(EXCEL_FILE)
    ws = wb['All Contacts']
    headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column+1)]
    rows = []
    for r in range(4, ws.max_row+1):
        row = {headers[c-1]: ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)}
        if cv(row.get('GHL Contact ID')):
            rows.append(row)
    return rows

def set_type(contact_id, api_value, dry_run):
    if dry_run:
        return True
    try:
        resp = requests.put(
            f"{BASE_URL}/contacts/{contact_id}",
            json={'type': api_value},
            headers=HEADERS, timeout=10
        )
        if resp.status_code in [200, 201]:
            return True
        print(f"    ERROR {resp.status_code}: {resp.text[:200]}")
        return False
    except Exception as e:
        print(f"    EXCEPTION: {e}")
        return False

# ── MAIN ──────────────────────────────────────────────────────────────────────
def run(dry_run):
    rows      = read_contacts()
    to_update = [r for r in rows if cv(r.get('Needs Update','')) == 'YES']
    skipping  = len(rows) - len(to_update)

    change_counts = Counter(cv(r['Correct Type']) for r in to_update)

    print(f"\n{'[DRY RUN] ' if dry_run else ''}GHL CONTACT TYPE UPDATE")
    print(f"{'─'*65}")
    print(f"  Total contacts  : {len(rows)}")
    print(f"  Updating        : {len(to_update)}")
    print(f"  Already correct : {skipping} (skipped)")
    print(f"\n  Changes by type:")
    for t in ['Active Customer','Past Customer','Lead','Internal','Vendor','Do Not Contact']:
        if change_counts[t]:
            print(f"    {t:20} : {change_counts[t]}")
    print(f"{'─'*65}\n")

    ok = err = 0
    for i, row in enumerate(to_update, 1):
        contact_id   = cv(row['GHL Contact ID'])
        correct_type = cv(row['Correct Type'])
        current      = cv(row['Current GHL Type'])
        api_value    = TYPE_MAP.get(correct_type)
        name = f"{cv(row['First Name'])} {cv(row['Last Name'])}".strip() or contact_id[:18]

        if not api_value:
            print(f"  [{i:3}] SKIP   {name} — unknown type '{correct_type}'")
            continue

        print(f"  [{i:3}] {name:32} {current:20} → {correct_type}")

        success = set_type(contact_id, api_value, dry_run)
        if success: ok += 1
        else: err += 1
        time.sleep(DELAY_SECONDS)

    print(f"\n{'─'*65}")
    print(f"{'[DRY RUN] ' if dry_run else ''}COMPLETE")
    print(f"  ✅ Success : {ok}")
    print(f"  ❌ Errors  : {err}")
    if dry_run:
        print(f"\n  Looks right? Apply for real with:")
        print(f"  python ghl_update.py --live")
    if err:
        print(f"\n  ⚠️  {err} errors — check output above.")

# ── ENTRY POINT ───────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if GHL_API_KEY == 'YOUR_API_KEY_HERE':
        print("❌ Set GHL_API_KEY in the script first."); sys.exit(1)

    live = '--live' in sys.argv
    if not live:
        print("DRY RUN — no changes will be made. Add --live to apply.\n")
        run(dry_run=True)
    else:
        confirm = input("⚠️  LIVE MODE — updates contacts in GHL. Type YES to confirm: ")
        if confirm.strip() != 'YES':
            print("Aborted."); sys.exit()
        run(dry_run=False)
