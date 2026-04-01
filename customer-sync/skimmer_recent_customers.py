#!/usr/bin/env python3
"""
Build a read-only report of active Skimmer customers whose matched GHL contact
was created within the last 6 months.

Usage:
    python skimmer_recent_customers.py
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path
import re

import requests
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook

SKIMMER_FILE = "Skimmer-Full-Customers.xlsx"
REPORT_FILE = "Skimmer_Active_Customers_6_Months_or_Less.xlsx"
PAGE_LIMIT = 100


def load_dotenv(path: str = ".env") -> dict[str, str]:
    env: dict[str, str] = {}
    env_path = Path(path)
    if not env_path.exists():
        raise FileNotFoundError(f"Missing {path}")
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "none", "nan"} else text


def normalize_phone(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D", "", text)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits if len(digits) == 10 else ""


def normalize_email(value: object) -> str:
    text = clean_text(value)
    return text.lower() if text else ""


def normalize_name(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    return re.sub(r"\s+", " ", text)


def parse_ghl_datetime(value: str) -> datetime | None:
    text = clean_text(value)
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def build_session(env: dict[str, str]) -> tuple[requests.Session, str]:
    required = ["GHL_BASE_URL", "GHL_LOCATION_ID", "GHL_TOKEN"]
    missing = [key for key in required if not clean_text(env.get(key))]
    if missing:
        raise ValueError(f"Missing required .env keys: {', '.join(missing)}")

    base_url = env["GHL_BASE_URL"].rstrip("/")
    version = clean_text(env.get("GHL_VERSION")) or "2021-07-28"
    session = requests.Session()
    session.headers.update(
        {
            "Authorization": f"Bearer {env['GHL_TOKEN']}",
            "Content-Type": "application/json",
            "Version": version,
        }
    )
    return session, base_url


def fetch_all_ghl_contacts(
    session: requests.Session, base_url: str, location_id: str
) -> list[dict]:
    contacts: list[dict] = []
    search_after: list[object] | None = None

    while True:
        payload: dict[str, object] = {"locationId": location_id, "pageLimit": PAGE_LIMIT}
        if search_after:
            payload["searchAfter"] = search_after
        response = session.post(f"{base_url}/contacts/search", json=payload, timeout=30)
        response.raise_for_status()
        page = response.json().get("contacts", [])
        if not page:
            break
        contacts.extend(page)
        search_after = page[-1].get("searchAfter")
        if len(page) < PAGE_LIMIT or not search_after:
            break

    return contacts


def load_active_skimmer_rows(path: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
    index = {header: idx for idx, header in enumerate(headers)}

    rows: list[dict] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        status = clean_text(values[index["Status"]])
        if status != "Active":
            continue

        first_name = clean_text(values[index["FirstName"]])
        last_name = clean_text(values[index["LastName"]])
        company_name = clean_text(values[index["CompanyName"]])
        full_name = clean_text(values[index["FullName"]])
        display_name = clean_text(values[index["FullNameOrCompanyDisplay"]])
        rows.append(
            {
                "status": status,
                "full_name": full_name,
                "display_name": display_name,
                "company_name": company_name,
                "phones": [
                    normalize_phone(values[index["MobilePhone1"]]),
                    normalize_phone(values[index["MobilePhone2"]]),
                    normalize_phone(values[index["HomePhone"]]),
                    normalize_phone(values[index["WorkPhone"]]),
                ],
                "emails": [
                    normalize_email(values[index["Email1"]]),
                    normalize_email(values[index["Email2"]]),
                    normalize_email(values[index["Email3"]]),
                    normalize_email(values[index["Email4"]]),
                ],
                "names": [
                    normalize_name(full_name),
                    normalize_name(display_name),
                    normalize_name(f"{first_name} {last_name}"),
                    normalize_name(company_name),
                ],
            }
        )

    for row in rows:
        row["phones"] = [phone for phone in row["phones"] if phone]
        row["emails"] = [email for email in row["emails"] if email]
        row["names"] = sorted({name for name in row["names"] if name})

    return rows


def build_ghl_indexes(contacts: list[dict]) -> dict[str, dict[str, list[dict]]]:
    by_phone: defaultdict[str, list[dict]] = defaultdict(list)
    by_email: defaultdict[str, list[dict]] = defaultdict(list)
    by_name: defaultdict[str, list[dict]] = defaultdict(list)

    for contact in contacts:
        phones = [normalize_phone(contact.get("phone"))]
        phones.extend(normalize_phone(value) for value in (contact.get("additionalPhones") or []))
        emails = [normalize_email(contact.get("email"))]
        emails.extend(normalize_email(value) for value in (contact.get("additionalEmails") or []))
        names = [
            normalize_name(contact.get("contactName")),
            normalize_name(
                f"{clean_text(contact.get('firstName'))} {clean_text(contact.get('lastName'))}"
            ),
            normalize_name(contact.get("companyName")),
        ]

        for phone in [value for value in phones if value]:
            by_phone[phone].append(contact)
        for email in [value for value in emails if value]:
            by_email[email].append(contact)
        for name in [value for value in names if value]:
            by_name[name].append(contact)

    return {"phone": dict(by_phone), "email": dict(by_email), "name": dict(by_name)}


def match_candidates(skimmer_row: dict, indexes: dict[str, dict[str, list[dict]]]) -> list[dict]:
    candidate_map: dict[str, dict] = {}

    for phone in skimmer_row["phones"]:
        for contact in indexes["phone"].get(phone, []):
            entry = candidate_map.setdefault(contact["id"], {"contact": contact, "methods": set()})
            entry["methods"].add("phone")
    for email in skimmer_row["emails"]:
        for contact in indexes["email"].get(email, []):
            entry = candidate_map.setdefault(contact["id"], {"contact": contact, "methods": set()})
            entry["methods"].add("email")
    for name in skimmer_row["names"]:
        for contact in indexes["name"].get(name, []):
            entry = candidate_map.setdefault(contact["id"], {"contact": contact, "methods": set()})
            entry["methods"].add("name")

    candidates = list(candidate_map.values())
    for candidate in candidates:
        candidate["methods"] = sorted(candidate["methods"])
    return candidates


def display_skimmer_name(row: dict) -> str:
    return row["display_name"] or row["full_name"] or row["company_name"] or "(unnamed)"


def display_ghl_name(contact: dict) -> str:
    return (
        clean_text(contact.get("contactName"))
        or " ".join(
            part
            for part in [clean_text(contact.get("firstName")), clean_text(contact.get("lastName"))]
            if part
        )
        or clean_text(contact.get("email"))
        or clean_text(contact.get("phone"))
        or clean_text(contact.get("id"))
    )


def candidate_summary(candidates: list[dict]) -> str:
    parts = []
    for candidate in candidates:
        contact = candidate["contact"]
        created_at = parse_ghl_datetime(clean_text(contact.get("dateAdded")))
        created_label = created_at.date().isoformat() if created_at else "unknown"
        parts.append(
            f"{display_ghl_name(contact)} "
            f"[{clean_text(contact.get('type')) or '(blank)'} | created {created_label}] "
            f"via {','.join(candidate['methods'])}"
        )
    return " | ".join(parts)


def classify_active_customers(active_rows: list[dict], indexes: dict[str, dict[str, list[dict]]]) -> dict[str, list[dict]]:
    buckets = {"recent": [], "older": [], "review": []}
    cutoff_date = date.today() - relativedelta(months=6)

    for row in active_rows:
        candidates = match_candidates(row, indexes)
        if not candidates:
            buckets["review"].append(
                {
                    "skimmer_name": display_skimmer_name(row),
                    "reason": "no_ghl_match",
                    "candidate_count": 0,
                    "has_recent_candidate": "NO",
                    "candidate_creation_dates": "",
                    "candidate_details": "",
                }
            )
            continue

        if len(candidates) > 1:
            candidate_dates = []
            has_recent_candidate = False
            for candidate in candidates:
                created_at = parse_ghl_datetime(clean_text(candidate["contact"].get("dateAdded")))
                if created_at:
                    candidate_dates.append(created_at.date().isoformat())
                    if created_at.date() >= cutoff_date:
                        has_recent_candidate = True
                else:
                    candidate_dates.append("unknown")
            buckets["review"].append(
                {
                    "skimmer_name": display_skimmer_name(row),
                    "reason": "multiple_ghl_matches",
                    "candidate_count": len(candidates),
                    "has_recent_candidate": "YES" if has_recent_candidate else "NO",
                    "candidate_creation_dates": " | ".join(candidate_dates),
                    "candidate_details": candidate_summary(candidates),
                }
            )
            continue

        contact = candidates[0]["contact"]
        created_at = parse_ghl_datetime(clean_text(contact.get("dateAdded")))
        if not created_at:
            buckets["review"].append(
                {
                    "skimmer_name": display_skimmer_name(row),
                    "reason": "missing_ghl_creation_date",
                    "candidate_count": 1,
                    "has_recent_candidate": "unknown",
                    "candidate_creation_dates": "unknown",
                    "candidate_details": candidate_summary(candidates),
                }
            )
            continue

        record = {
            "skimmer_name": display_skimmer_name(row),
            "ghl_contact_id": clean_text(contact.get("id")),
            "ghl_name": display_ghl_name(contact),
            "ghl_type": clean_text(contact.get("type")),
            "ghl_phone": clean_text(contact.get("phone")),
            "ghl_email": clean_text(contact.get("email")),
            "created_at": created_at,
            "created_date": created_at.date(),
            "days_old": (date.today() - created_at.date()).days,
            "match_methods": ",".join(candidates[0]["methods"]),
        }

        if record["created_date"] >= cutoff_date:
            buckets["recent"].append(record)
        else:
            buckets["older"].append(record)

    return buckets


def autosize(sheet) -> None:
    for column_cells in sheet.columns:
        width = 0
        letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, len(value))
        sheet.column_dimensions[letter].width = min(width + 2, 60)


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    autosize(sheet)


def write_report(report_path: str, active_rows: list[dict], buckets: dict[str, list[dict]]) -> None:
    cutoff_date = date.today() - relativedelta(months=6)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    summary_rows = [
        ["Metric", "Value", "Notes"],
        ["Report date", date.today().isoformat(), ""],
        ["6 month cutoff", cutoff_date.isoformat(), "Customers created on or after this date are included"],
        ["Active Skimmer customers", len(active_rows), ""],
        ["Matched recent customers", len(buckets["recent"]), "Created on or after cutoff in GHL"],
        ["Matched older customers", len(buckets["older"]), "Created before cutoff in GHL"],
        ["Review needed", len(buckets["review"]), "No match or multiple GHL matches"],
        ["", "", ""],
        ["Recent customers by GHL type", "", ""],
    ]

    for ghl_type, count in sorted(Counter(row["ghl_type"] or "(blank)" for row in buckets["recent"]).items()):
        summary_rows.append([ghl_type, count, ""])

    for row in summary_rows:
        summary.append(row)
    autosize(summary)

    add_sheet(
        workbook,
        "Recent Active Customers",
        [
            "Skimmer Name",
            "GHL Contact ID",
            "GHL Name",
            "GHL Type",
            "GHL Creation Date",
            "Days Since Created",
            "Phone",
            "Email",
            "Match Methods",
        ],
        [
            [
                row["skimmer_name"],
                row["ghl_contact_id"],
                row["ghl_name"],
                row["ghl_type"],
                row["created_date"].isoformat(),
                row["days_old"],
                row["ghl_phone"],
                row["ghl_email"],
                row["match_methods"],
            ]
            for row in sorted(buckets["recent"], key=lambda item: (item["created_date"], item["skimmer_name"]))
        ],
    )

    add_sheet(
        workbook,
        "Older Active Customers",
        [
            "Skimmer Name",
            "GHL Contact ID",
            "GHL Name",
            "GHL Type",
            "GHL Creation Date",
            "Days Since Created",
            "Phone",
            "Email",
            "Match Methods",
        ],
        [
            [
                row["skimmer_name"],
                row["ghl_contact_id"],
                row["ghl_name"],
                row["ghl_type"],
                row["created_date"].isoformat(),
                row["days_old"],
                row["ghl_phone"],
                row["ghl_email"],
                row["match_methods"],
            ]
            for row in sorted(buckets["older"], key=lambda item: (item["created_date"], item["skimmer_name"]))
        ],
    )

    add_sheet(
        workbook,
        "Review Needed",
        [
            "Skimmer Name",
            "Reason",
            "Candidate Count",
            "Has Recent Candidate",
            "Candidate Creation Dates",
            "Candidate Details",
        ],
        [
            [
                row["skimmer_name"],
                row["reason"],
                row["candidate_count"],
                row["has_recent_candidate"],
                row["candidate_creation_dates"],
                row["candidate_details"],
            ]
            for row in sorted(buckets["review"], key=lambda item: item["skimmer_name"])
        ],
    )

    workbook.save(report_path)


def print_summary(active_rows: list[dict], buckets: dict[str, list[dict]]) -> None:
    cutoff_date = date.today() - relativedelta(months=6)
    print()
    print("ACTIVE SKIMMER CUSTOMERS CREATED WITHIN 6 MONTHS")
    print("-" * 64)
    print(f"Report date          : {date.today().isoformat()}")
    print(f"6 month cutoff       : {cutoff_date.isoformat()}")
    print(f"Active Skimmer rows  : {len(active_rows)}")
    print(f"Recent matches       : {len(buckets['recent'])}")
    print(f"Older matches        : {len(buckets['older'])}")
    print(f"Review needed        : {len(buckets['review'])}")
    print("-" * 64)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--skimmer-file", default=SKIMMER_FILE)
    parser.add_argument("--report-file", default=REPORT_FILE)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        env = load_dotenv()
        session, base_url = build_session(env)
        contacts = fetch_all_ghl_contacts(session, base_url, env["GHL_LOCATION_ID"])
        active_rows = load_active_skimmer_rows(args.skimmer_file)
        indexes = build_ghl_indexes(contacts)
        buckets = classify_active_customers(active_rows, indexes)
        write_report(args.report_file, active_rows, buckets)
        print_summary(active_rows, buckets)
        print(f"Report written to: {args.report_file}")
        return 0
    except Exception as exc:
        print(f"Report failed: {exc}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
