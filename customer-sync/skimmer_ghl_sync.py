#!/usr/bin/env python3
"""
Sync GHL contact types from the current Skimmer customer export.

Default mode is a dry run:
    python skimmer_ghl_sync.py

Write changes to GHL only when explicitly requested:
    python skimmer_ghl_sync.py --live

Required .env keys:
    GHL_BASE_URL
    GHL_VERSION
    GHL_LOCATION_ID
    GHL_TOKEN
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from datetime import datetime

import requests
from openpyxl import Workbook, load_workbook

SKIMMER_FILE = "Skimmer-Full-Customers.xlsx"
REPORT_FILE = "Skimmer_GHL_Sync_Review.xlsx"
PAGE_LIMIT = 100
LIVE_DELAY_SECONDS = 0.25

STATUS_TO_TYPE = {
    "Active": "customer",
    "Inactive": "past_customer",
}

TYPE_LABELS = {
    "customer": "Active Customer",
    "past_customer": "Past Customer",
    "lead": "Lead",
    "internal": "Internal",
    "vendor": "Vendor",
    "do_not_contact": "Do Not Contact",
    "": "(blank)",
}

PROTECTED_CURRENT_TYPES = {"internal", "vendor", "do_not_contact"}


def load_dotenv(path: str = ".env") -> dict[str, str]:
    env: dict[str, str] = {}
    env_path = Path(path)
    if not env_path.exists():
        raise FileNotFoundError(f"Missing {path}")

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"", "none", "nan"} else text


def normalize_phone(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    digits = re.sub(r"\D", "", text)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    return digits if len(digits) == 10 else ""


def normalize_email(value: object) -> str:
    text = clean_text(value)
    return text.lower() if text else ""


def normalize_name(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    return re.sub(r"\s+", " ", text)


def pretty_type(type_value: str) -> str:
    return TYPE_LABELS.get(type_value or "", type_value or "(blank)")


def build_session(env: dict[str, str]) -> tuple[requests.Session, str]:
    required = ["GHL_BASE_URL", "GHL_LOCATION_ID", "GHL_TOKEN"]
    missing = [key for key in required if not clean_text(env.get(key))]
    if missing:
        raise ValueError(f"Missing required .env keys: {', '.join(missing)}")

    base_url = env["GHL_BASE_URL"].rstrip("/")
    version = clean_text(env.get("GHL_VERSION")) or "2021-07-28"

    session = requests.Session()
    session.headers.update(
        {
            "Authorization": f"Bearer {env['GHL_TOKEN']}",
            "Content-Type": "application/json",
            "Version": version,
        }
    )
    return session, base_url


def fetch_all_ghl_contacts(
    session: requests.Session, base_url: str, location_id: str
) -> list[dict]:
    contacts: list[dict] = []
    search_after: list[object] | None = None

    while True:
        payload: dict[str, object] = {
            "locationId": location_id,
            "pageLimit": PAGE_LIMIT,
        }
        if search_after:
            payload["searchAfter"] = search_after

        response = session.post(
            f"{base_url}/contacts/search", json=payload, timeout=30
        )
        response.raise_for_status()
        page = response.json().get("contacts", [])
        if not page:
            break

        contacts.extend(page)
        search_after = page[-1].get("searchAfter")
        if len(page) < PAGE_LIMIT or not search_after:
            break

    return contacts


def load_skimmer_rows(path: str) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
    index = {header: idx for idx, header in enumerate(headers)}

    rows: list[dict] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        status = clean_text(values[index["Status"]])
        if status not in STATUS_TO_TYPE:
            continue

        first_name = clean_text(values[index["FirstName"]])
        last_name = clean_text(values[index["LastName"]])
        company_name = clean_text(values[index["CompanyName"]])
        full_name = clean_text(values[index["FullName"]])
        display_name = clean_text(values[index["FullNameOrCompanyDisplay"]])

        row = {
            "status": status,
            "desired_type": STATUS_TO_TYPE[status],
            "full_name": full_name,
            "display_name": display_name,
            "company_name": company_name,
            "phones": [
                normalize_phone(values[index["MobilePhone1"]]),
                normalize_phone(values[index["MobilePhone2"]]),
                normalize_phone(values[index["HomePhone"]]),
                normalize_phone(values[index["WorkPhone"]]),
            ],
            "emails": [
                normalize_email(values[index["Email1"]]),
                normalize_email(values[index["Email2"]]),
                normalize_email(values[index["Email3"]]),
                normalize_email(values[index["Email4"]]),
            ],
            "names": [
                normalize_name(full_name),
                normalize_name(display_name),
                normalize_name(f"{first_name} {last_name}"),
                normalize_name(company_name),
            ],
        }

        row["phones"] = [phone for phone in row["phones"] if phone]
        row["emails"] = [email for email in row["emails"] if email]
        row["names"] = sorted({name for name in row["names"] if name})
        rows.append(row)

    return rows


def build_skimmer_indexes(skimmer_rows: list[dict]) -> dict[str, dict[str, list[dict]]]:
    by_phone: defaultdict[str, list[dict]] = defaultdict(list)
    by_email: defaultdict[str, list[dict]] = defaultdict(list)
    by_name: defaultdict[str, list[dict]] = defaultdict(list)

    for row in skimmer_rows:
        for phone in row["phones"]:
            by_phone[phone].append(row)
        for email in row["emails"]:
            by_email[email].append(row)
        for name in row["names"]:
            by_name[name].append(row)

    return {
        "phone": dict(by_phone),
        "email": dict(by_email),
        "name": dict(by_name),
    }


def contact_match_candidates(contact: dict, indexes: dict[str, dict[str, list[dict]]]) -> list[dict]:
    candidate_map: dict[int, dict] = {}

    phones = [normalize_phone(contact.get("phone"))]
    phones.extend(normalize_phone(value) for value in (contact.get("additionalPhones") or []))
    for phone in [value for value in phones if value]:
        for skimmer_row in indexes["phone"].get(phone, []):
            entry = candidate_map.setdefault(
                id(skimmer_row), {"record": skimmer_row, "methods": set()}
            )
            entry["methods"].add("phone")

    emails = [normalize_email(contact.get("email"))]
    emails.extend(normalize_email(value) for value in (contact.get("additionalEmails") or []))
    for email in [value for value in emails if value]:
        for skimmer_row in indexes["email"].get(email, []):
            entry = candidate_map.setdefault(
                id(skimmer_row), {"record": skimmer_row, "methods": set()}
            )
            entry["methods"].add("email")

    names = [
        normalize_name(contact.get("contactName")),
        normalize_name(
            f"{clean_text(contact.get('firstName'))} {clean_text(contact.get('lastName'))}"
        ),
        normalize_name(contact.get("companyName")),
    ]
    for name in [value for value in names if value]:
        for skimmer_row in indexes["name"].get(name, []):
            entry = candidate_map.setdefault(
                id(skimmer_row), {"record": skimmer_row, "methods": set()}
            )
            entry["methods"].add("name")

    candidates = list(candidate_map.values())
    for candidate in candidates:
        candidate["methods"] = sorted(candidate["methods"])
    return candidates


def decide_contact(contact: dict, candidates: list[dict]) -> dict:
    current_type = clean_text(contact.get("type"))

    if not candidates:
        return {
            "decision": "review",
            "reason": "no_skimmer_match",
            "desired_type": "",
        }

    statuses = {candidate["record"]["status"] for candidate in candidates}
    if len(candidates) == 1:
        desired_type = candidates[0]["record"]["desired_type"]
        reason = "unique_match"
    elif len(statuses) == 1:
        desired_type = STATUS_TO_TYPE[next(iter(statuses))]
        reason = "multiple_matches_same_status"
    else:
        return {
            "decision": "review",
            "reason": "ambiguous_mixed_status",
            "desired_type": "",
        }

    if current_type in PROTECTED_CURRENT_TYPES and current_type != desired_type:
        return {
            "decision": "review",
            "reason": "protected_current_type",
            "desired_type": desired_type,
        }

    if current_type == desired_type:
        return {
            "decision": "already_correct",
            "reason": reason,
            "desired_type": desired_type,
        }

    return {
        "decision": "update",
        "reason": reason,
        "desired_type": desired_type,
    }


def row_name(contact: dict) -> str:
    return (
        clean_text(contact.get("contactName"))
        or " ".join(
            part
            for part in [clean_text(contact.get("firstName")), clean_text(contact.get("lastName"))]
            if part
        )
        or clean_text(contact.get("email"))
        or clean_text(contact.get("phone"))
        or clean_text(contact.get("id"))
    )


def candidate_summary(candidates: list[dict]) -> str:
    parts = []
    for candidate in candidates:
        record = candidate["record"]
        label = record["full_name"] or record["display_name"] or record["company_name"] or "(unnamed)"
        methods = ",".join(candidate["methods"])
        parts.append(f"{label} [{record['status']}] via {methods}")
    return " | ".join(parts)


def classify_contacts(contacts: list[dict], indexes: dict[str, dict[str, list[dict]]]) -> dict[str, list[dict]]:
    buckets = {
        "update": [],
        "already_correct": [],
        "review": [],
    }

    for contact in contacts:
        candidates = contact_match_candidates(contact, indexes)
        decision = decide_contact(contact, candidates)
        record = {
            "contact_id": clean_text(contact.get("id")),
            "name": row_name(contact),
            "current_type": clean_text(contact.get("type")),
            "desired_type": decision["desired_type"],
            "reason": decision["reason"],
            "phone": clean_text(contact.get("phone")),
            "email": clean_text(contact.get("email")),
            "match_count": len(candidates),
            "match_details": candidate_summary(candidates),
            "candidates": candidates,
            "contact": contact,
        }
        buckets[decision["decision"]].append(record)

    return buckets


def autosize(sheet) -> None:
    for column_cells in sheet.columns:
        width = 0
        letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, len(value))
        sheet.column_dimensions[letter].width = min(width + 2, 60)


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    autosize(sheet)


def write_report(
    report_path: str,
    contacts: list[dict],
    skimmer_rows: list[dict],
    buckets: dict[str, list[dict]],
) -> str:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    summary_rows = [
        ["Metric", "Value", "Notes"],
        ["GHL contacts fetched", len(contacts), "Live data pulled from GHL"],
        ["Skimmer customers loaded", len(skimmer_rows), "Rows with Active/Inactive status"],
        ["Skimmer Active", sum(1 for row in skimmer_rows if row["status"] == "Active"), "Maps to Active Customer"],
        ["Skimmer Inactive", sum(1 for row in skimmer_rows if row["status"] == "Inactive"), "Maps to Past Customer"],
        ["Safe updates", len(buckets["update"]), "Dry-run changes ready to apply"],
        ["Already correct", len(buckets["already_correct"]), "Matched and already aligned"],
        ["Review needed", len(buckets["review"]), "No match, protected type, or ambiguity"],
        ["", "", ""],
        ["Update breakdown", "", ""],
    ]

    for (current_type, desired_type), count in sorted(
        Counter((row["current_type"], row["desired_type"]) for row in buckets["update"]).items()
    ):
        summary_rows.append(
            [f"{pretty_type(current_type)} -> {pretty_type(desired_type)}", count, ""]
        )

    summary_rows.append(["", "", ""])
    summary_rows.append(["Review breakdown", "", ""])
    for reason, count in sorted(Counter(row["reason"] for row in buckets["review"]).items()):
        summary_rows.append([reason, count, ""])

    for row in summary_rows:
        summary.append(row)
    autosize(summary)

    add_sheet(
        workbook,
        "Proposed Updates",
        [
            "GHL Contact ID",
            "Name",
            "Phone",
            "Email",
            "Current Type",
            "Desired Type",
            "Reason",
            "Match Count",
            "Match Details",
        ],
        [
            [
                row["contact_id"],
                row["name"],
                row["phone"],
                row["email"],
                pretty_type(row["current_type"]),
                pretty_type(row["desired_type"]),
                row["reason"],
                row["match_count"],
                row["match_details"],
            ]
            for row in buckets["update"]
        ],
    )

    add_sheet(
        workbook,
        "Already Correct",
        [
            "GHL Contact ID",
            "Name",
            "Phone",
            "Email",
            "Type",
            "Reason",
            "Match Count",
            "Match Details",
        ],
        [
            [
                row["contact_id"],
                row["name"],
                row["phone"],
                row["email"],
                pretty_type(row["current_type"]),
                row["reason"],
                row["match_count"],
                row["match_details"],
            ]
            for row in buckets["already_correct"]
        ],
    )

    add_sheet(
        workbook,
        "Review Needed",
        [
            "GHL Contact ID",
            "Name",
            "Phone",
            "Email",
            "Current Type",
            "Suggested Type",
            "Reason",
            "Match Count",
            "Match Details",
        ],
        [
            [
                row["contact_id"],
                row["name"],
                row["phone"],
                row["email"],
                pretty_type(row["current_type"]),
                pretty_type(row["desired_type"]),
                row["reason"],
                row["match_count"],
                row["match_details"],
            ]
            for row in buckets["review"]
        ],
    )

    try:
        workbook.save(report_path)
        return report_path
    except PermissionError:
        path = Path(report_path)
        fallback = path.with_name(
            f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
        )
        workbook.save(fallback)
        return str(fallback)


def print_summary(contacts: list[dict], skimmer_rows: list[dict], buckets: dict[str, list[dict]], dry_run: bool) -> None:
    print()
    print("SKIMMER -> GHL CONTACT TYPE SYNC")
    print("-" * 64)
    print(f"Mode              : {'DRY RUN' if dry_run else 'LIVE'}")
    print(f"GHL contacts      : {len(contacts)}")
    print(f"Skimmer customers : {len(skimmer_rows)}")
    print(f"Safe updates      : {len(buckets['update'])}")
    print(f"Already correct   : {len(buckets['already_correct'])}")
    print(f"Review needed     : {len(buckets['review'])}")

    if buckets["update"]:
        print()
        print("Updates by type:")
        for (current_type, desired_type), count in sorted(
            Counter((row["current_type"], row["desired_type"]) for row in buckets["update"]).items()
        ):
            print(
                f"  {pretty_type(current_type):16} -> {pretty_type(desired_type):16} : {count}"
            )

    if buckets["review"]:
        print()
        print("Review reasons:")
        for reason, count in sorted(Counter(row["reason"] for row in buckets["review"]).items()):
            print(f"  {reason:24} : {count}")

    print("-" * 64)


def apply_updates(
    session: requests.Session, base_url: str, rows: list[dict], dry_run: bool
) -> tuple[int, int]:
    success = 0
    errors = 0

    for index, row in enumerate(rows, start=1):
        print(
            f"[{index:3}] {row['name'][:32]:32} "
            f"{pretty_type(row['current_type']):16} -> {pretty_type(row['desired_type'])}"
        )

        if dry_run:
            success += 1
            continue

        response = session.put(
            f"{base_url}/contacts/{row['contact_id']}",
            json={"type": row["desired_type"]},
            timeout=20,
        )
        if response.status_code in {200, 201}:
            success += 1
        else:
            errors += 1
            print(f"      ERROR {response.status_code}: {response.text[:200]}")
        time.sleep(LIVE_DELAY_SECONDS)

    return success, errors


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--live", action="store_true", help="Apply updates to GHL")
    parser.add_argument("--skimmer-file", default=SKIMMER_FILE)
    parser.add_argument("--report-file", default=REPORT_FILE)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    dry_run = not args.live

    try:
        env = load_dotenv()
        session, base_url = build_session(env)
        contacts = fetch_all_ghl_contacts(session, base_url, env["GHL_LOCATION_ID"])
        skimmer_rows = load_skimmer_rows(args.skimmer_file)
    except Exception as exc:
        print(f"Setup failed: {exc}")
        return 1

    indexes = build_skimmer_indexes(skimmer_rows)
    buckets = classify_contacts(contacts, indexes)
    report_path = write_report(args.report_file, contacts, skimmer_rows, buckets)
    print_summary(contacts, skimmer_rows, buckets, dry_run)
    print(f"Report written to: {report_path}")

    if dry_run:
        print()
        print("Dry run only. No changes were sent to GHL.")
        print("Review the workbook, then rerun with --live if it looks right.")
        return 0

    print()
    confirm = input("LIVE MODE - this will update GHL contact types. Type YES to continue: ")
    if confirm.strip() != "YES":
        print("Aborted.")
        return 1

    success, errors = apply_updates(session, base_url, buckets["update"], dry_run=False)
    print()
    print("LIVE RUN COMPLETE")
    print(f"  Success : {success}")
    print(f"  Errors  : {errors}")
    return 0 if errors == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
